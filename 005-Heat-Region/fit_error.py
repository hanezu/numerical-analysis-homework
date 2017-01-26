import openpyxl

import numpy as np
import scipy.optimize
import pandas as pd
from matplotlib.ticker import LinearLocator, FormatStrFormatter

from matplotlib import cm
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D


# data which you want to fit
# wb = openpyxl.load_workbook('error_fit.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
# N_list = []
# error_list = []
# for row in range(2,sheet.max_row):
#     N_list.append(sheet.cell(row=row,column=1).value)
#     error_list.append(sheet.cell(row=row,column=2).value)

df_error = pd.read_csv("error_fit.csv", delimiter=' ')
normal_error = 1
df_error = df_error[df_error['error'] < normal_error]

xdata = df_error['N']
ydata = df_error['r']
zdata = 1/df_error['error']
# initial guess for the parameters
parameter_initial = np.array([0.0, 0.0])  # a, b, c
xdata = xdata.reshape(10,10).T
ydata = ydata.reshape(10,10).T
zdata = zdata.reshape(10,10).T

# function to fit
def func(x, a, b):
    return (a + b * x)


fig = plt.figure()
# ax = fig.add_subplot(111, projection='3d')
ax = fig.add_subplot(111, projection='3d')
# ax.scatter(xdata, ydata, zdata, c='r', marker='x')
surf = ax.plot_surface(xdata, ydata, zdata, rstride=1, cstride=1, cmap=cm.coolwarm,
                       linewidth=0, antialiased=False)

ax.set_xlabel('N')
ax.set_ylabel('r')
ax.set_zlabel('1/error')
ax.zaxis.set_major_formatter(FormatStrFormatter('%.1e'))
fig.colorbar(surf, shrink=0.5, aspect=5)


plt.show()


# paramater_optimal, covariance = scipy.optimize.curve_fit(func, xdata, ydata, p0=parameter_initial)
# print("paramater =", paramater_optimal)
#
# y = func(xdata, paramater_optimal[0], paramater_optimal[1])
# plt.plot(xdata, ydata, 'o')
# plt.plot(xdata, y, '-')
# plt.ylabel('1/error')
# plt.xlabel('N')
#
# plt.show()
